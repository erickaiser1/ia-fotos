import os
import datetime
import pandas as pd
import cv2
import unicodedata
from ultralytics import YOLO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# -----------------------------
# CONFIG
# -----------------------------

PLANILHA = os.path.join("planilha", "prioridades.xlsx")
MODELO = r"C:\Users\erick.brandao\Downloads\best.pt"

OUTPUT_FOLDER = "outputs"
PROCESSADAS = "processadas"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(PROCESSADAS, exist_ok=True)

# -----------------------------
# NORMALIZAR TEXTO
# -----------------------------

def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto

# -----------------------------
# CARREGAR MODELO
# -----------------------------

print("Carregando modelo YOLO...")
model = YOLO(MODELO)
print("Modelo carregado!")

# -----------------------------
# FUNÇÃO PRINCIPAL
# -----------------------------

def rodar_analise(pasta_fotos):

    agora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    RELATORIO = os.path.join(OUTPUT_FOLDER, f"relatorio_{agora}.xlsx")

    # -----------------------------
    # CARREGAR PLANILHA
    # -----------------------------

    try:
        df = pd.read_excel(PLANILHA)
    except:
        df = pd.DataFrame(columns=["PROBLEMAS", "PRIORIDADE"])

    coluna_problema = "PROBLEMAS"
    coluna_prioridade = "PRIORIDADE"

    if coluna_problema not in df.columns:
        df[coluna_problema] = ""

    if coluna_prioridade not in df.columns:
        df[coluna_prioridade] = 0

    df["PROBLEMA_NORMALIZADO"] = df[coluna_problema].apply(normalizar)

    # -----------------------------
    # CLASSIFICAÇÃO
    # -----------------------------

    def classificar_imagem(caminho_img):

        results = model(caminho_img)
        probs = results[0].probs

        if probs is None:
            return "Sem problemas", 0

        classe = results[0].names[probs.top1]
        confianca = float(probs.top1conf)

        if confianca < 0.5:
            return "Sem problemas", confianca

        return classe, confianca

    # -----------------------------
    # BUSCAR IMAGENS
    # -----------------------------

    arquivos_imagem = []

    for raiz, _, arquivos in os.walk(pasta_fotos):
        for arquivo in arquivos:
            if arquivo.lower().endswith((".jpg",".jpeg",".png")):
                arquivos_imagem.append(os.path.join(raiz,arquivo))

    print(f"Total de imagens: {len(arquivos_imagem)}")

    # -----------------------------
    # PROCESSAMENTO
    # -----------------------------

    resultados = []

    for caminho in arquivos_imagem:

        nome = os.path.basename(caminho)
        print("Analisando:", nome)

        imagem = cv2.imread(caminho)

        if imagem is None:
            continue

        categoria, conf = classificar_imagem(caminho)

        # Desenhar na imagem
        cv2.putText(imagem, f"{categoria} ({conf:.2f})",
                    (20,40),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,(0,0,255),2)

        caminho_saida = os.path.join(PROCESSADAS, nome)
        cv2.imwrite(caminho_saida, imagem)

        # PRIORIDADE
        if categoria == "Sem problemas":
            prioridade = 0
        else:
            categoria_norm = normalizar(categoria)
            prioridade = 0

            for _, row in df.iterrows():
                problema_norm = row["PROBLEMA_NORMALIZADO"]

                if categoria_norm in problema_norm or problema_norm in categoria_norm:
                    prioridade = row[coluna_prioridade]
                    break

        resultados.append({
            "Foto": nome,
            "Categoria": categoria,
            "Prioridade": prioridade,
            "Confiança": round(conf,3),
            "Imagem": caminho_saida
        })

    # -----------------------------
    # CRIAR EXCEL
    # -----------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    ws.append(["Foto","Categoria","Prioridade","Confiança","Imagem"])

    linha = 2

    for r in resultados:

        ws.cell(linha,1,r["Foto"])
        ws.cell(linha,2,r["Categoria"])
        ws.cell(linha,3,r["Prioridade"])
        ws.cell(linha,4,r["Confiança"])

        if os.path.exists(r["Imagem"]):
            img = ExcelImage(r["Imagem"])
            img.width = 150
            img.height = 150
            ws.add_image(img,f"E{linha}")
            ws.row_dimensions[linha].height = 120

        linha += 1

    ws.column_dimensions["E"].width = 40

    # -----------------------------
    # RESUMO SEGURO
    # -----------------------------

    df_result = pd.DataFrame(resultados)
    ws_resumo = wb.create_sheet("Resumo")

    # 🛑 sem dados
    if df_result.empty:

        ws_resumo.append(["Status", "Nenhuma imagem processada"])
        wb.save(RELATORIO)
        return RELATORIO

    # garantir coluna
    if "Prioridade" not in df_result.columns:
        df_result["Prioridade"] = 0

    # -----------------------------
    # MÉDIA PRIORIDADE
    # -----------------------------

    df_prioridade = pd.to_numeric(df_result["Prioridade"], errors='coerce')
    df_prioridade = df_prioridade[df_prioridade > 0]

    if df_prioridade.empty:
        media = 0
        status = "⚪ SEM DADOS"
    else:
        media = round(df_prioridade.mean(),2)

        if media <= 2:
            status = "🔴 ALTA PRIORIDADE"
        elif media <= 3:
            status = "🟡 ATENÇÃO"
        else:
            status = "🟢 TUDO CERTO"

    ws_resumo.append(["Média Prioridade", media, status])
    ws_resumo.append([])

    # -----------------------------
    # PROBLEMAS
    # -----------------------------

    df_filtrado = df_result[df_result["Categoria"] != "Sem problemas"]

    if df_filtrado.empty:

        ws_resumo.append(["Status", "Nenhum problema identificado"])

    else:

        contagem = df_filtrado["Categoria"].value_counts()
        total = contagem.sum()

        ws_resumo.append(["Problema","Qtd","%","Urgência"])

        for problema, qtd in contagem.items():

            perc = (qtd / total) * 100

            media_prob = pd.to_numeric(
                df_filtrado[df_filtrado["Categoria"] == problema]["Prioridade"],
                errors='coerce'
            ).mean()

            if pd.isna(media_prob):
                urgencia = "⚪ SEM DADO"
            elif media_prob <= 2:
                urgencia = "🔴 CRÍTICO"
            elif media_prob <= 3:
                urgencia = "🟡 MÉDIO"
            else:
                urgencia = "🟢 BAIXO"

            ws_resumo.append([problema, qtd, round(perc,1), urgencia])

    # -----------------------------
    # SALVAR
    # -----------------------------

    wb.save(RELATORIO)

    print("\n✅ Relatório criado:", RELATORIO)

    return RELATORIO